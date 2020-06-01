from datetime import datetime
from openpyxl import load_workbook
from classes import Product, Review
from mappings import *

# Using the read_only method since I'm not gonna be editing the spreadsheet
workbook = load_workbook(filename="reviews-sample.xlsx", read_only=True)
sheet = workbook.active

products = []
reviews = []

# Using values_only since I'm only interested in the cell value
for row in sheet.iter_rows(min_row=2, values_only=True):
  product = Product(id=row[PRODUCT_ID],
                    parent=row[PRODUCT_PARENT],
                    title=row[PRODUCT_TITLE],
                    category=row[PRODUCT_CATEGORY])
  products.append(product)

  # Need to parse the date from spreadsheet into a datetime format
  spread_date = row[REVIEW_DATE]
  parsed_date = datetime.strptime(spread_date, "%Y-%m-%d")

  review = Review(id=row[REVIEW_ID],
                  customer_id=row[REVIEW_CUSTOMER],
                  stars=row[REVIEW_STARS],
                  headline=row[REVIEW_HEADLINE],
                  body=row[REVIEW_BODY],
                  date=parsed_date)

  reviews.append(review)

print(products)
print(reviews)